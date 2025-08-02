from django.shortcuts import render
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from datetime import timedelta, datetime
from django_filters import rest_framework as filters

from .models import SalesAnalytics, CreditScore, LoanOffer, MonthlyReport, ProductPerformance, SalesTarget, PaymentAnalysis, ExportRequest
from .serializers import (
    SalesAnalyticsSerializer, CreditScoreSerializer, LoanOfferSerializer,
    MonthlyReportSerializer, ProductPerformanceSerializer,
    FarmerDashboardSerializer, SalesOverviewSerializer, CreditOverviewSerializer,
    ReportOverviewSerializer, SalesTargetSerializer, PaymentAnalysisSerializer, ExportRequestSerializer
)
from .services import AnalyticsCalculationService
from users.models import User
from orders.models import Order, OrderItem
from marketplace.models import Product


class SalesAnalyticsViewSet(viewsets.ModelViewSet):
    """ViewSet for sales analytics data"""
    serializer_class = SalesAnalyticsSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.DjangoFilterBackend]
    filterset_fields = ['date', 'farmer']
    ordering = ['-date']

    def get_queryset(self):
        """Return analytics for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return SalesAnalytics.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return SalesAnalytics.objects.all()
        return SalesAnalytics.objects.none()

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get comprehensive dashboard data for farmer"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            # Calculate sales analytics based on actual performance
            sales_analytics = AnalyticsCalculationService.calculate_sales_analytics(request.user)
            
            # Calculate credit score based on actual performance
            credit_score_obj = AnalyticsCalculationService.calculate_credit_score(request.user)
            
            # Get eligible loan offers based on actual performance
            loan_offers = AnalyticsCalculationService.get_eligible_loan_offers(request.user)
            
            # Get monthly reports
            monthly_reports = MonthlyReport.objects.filter(
                farmer=request.user
            ).order_by('-generated_at')[:2]
            
            # Extract data from calculated analytics
            total_revenue = sales_analytics.total_revenue if sales_analytics else 0
            total_sales = sales_analytics.total_sales if sales_analytics else 0
            avg_order_value = sales_analytics.avg_order_value if sales_analytics else 0
            revenue_growth = sales_analytics.revenue_growth if sales_analytics else 0
            sales_growth = sales_analytics.sales_growth if sales_analytics else 0
            top_products = sales_analytics.top_products if sales_analytics else []
            monthly_trends = sales_analytics.monthly_trends if sales_analytics else []

            dashboard_data = {
                'total_revenue': float(total_revenue),
                'total_sales': total_sales,
                'avg_order_value': float(avg_order_value),
                'revenue_growth': float(revenue_growth),
                'sales_growth': float(revenue_growth),  # Simplified for now
                'top_products': list(top_products),
                'monthly_trends': monthly_trends,
                'credit_score': credit_score_obj.score,
                'loan_offers': LoanOfferSerializer(loan_offers, many=True).data,
                'monthly_reports': MonthlyReportSerializer(monthly_reports, many=True).data
            }

            serializer = FarmerDashboardSerializer(data=dashboard_data)
            if serializer.is_valid():
                return Response(serializer.data)
            else:
                return Response(serializer.errors, status=400)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def sales_overview(self, request):
        """Get sales overview data"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            # Calculate sales metrics
            today = timezone.now().date()
            last_30_days = today - timedelta(days=30)
            last_month = today - timedelta(days=60)

            farmer_orders = Order.objects.filter(
                items__product__farmer=request.user
            ).distinct()

            total_revenue = farmer_orders.filter(
                status='delivered',
                created_at__gte=last_30_days
            ).aggregate(total=Sum('total_amount'))['total'] or 0

            total_sales = farmer_orders.filter(
                status='delivered',
                created_at__gte=last_30_days
            ).count()

            # Calculate growth
            current_month_revenue = farmer_orders.filter(
                status='delivered',
                created_at__gte=last_30_days
            ).aggregate(total=Sum('total_amount'))['total'] or 0

            previous_month_revenue = farmer_orders.filter(
                status='delivered',
                created_at__gte=last_month,
                created_at__lt=last_30_days
            ).aggregate(total=Sum('total_amount'))['total'] or 0

            revenue_growth = 0
            if previous_month_revenue > 0:
                revenue_growth = ((current_month_revenue - previous_month_revenue) / previous_month_revenue) * 100

            # Get top products
            top_products = OrderItem.objects.filter(
                order__in=farmer_orders,
                order__status='delivered',
                order__created_at__gte=last_30_days
            ).values('product__name').annotate(
                revenue=Sum(F('quantity') * F('unit_price')),
                sales=Sum('quantity')
            ).order_by('-revenue')[:4]

            # Find best month
            best_month = "Current Month"  # Simplified for now

            # Find top category
            top_category = "Vegetables"  # Simplified for now

            sales_data = {
                'total_revenue': float(total_revenue),
                'total_sales': total_sales,
                'revenue_growth': float(revenue_growth),
                'sales_growth': float(revenue_growth),
                'best_month': best_month,
                'top_category': top_category,
                'top_products': list(top_products)
            }

            serializer = SalesOverviewSerializer(data=sales_data)
            if serializer.is_valid():
                return Response(serializer.data)
            else:
                return Response(serializer.errors, status=400)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class CreditScoreViewSet(viewsets.ModelViewSet):
    """ViewSet for credit score management"""
    serializer_class = CreditScoreSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Return credit score for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return CreditScore.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return CreditScore.objects.all()
        return CreditScore.objects.none()

    @action(detail=False, methods=['get'])
    def overview(self, request):
        """Get credit overview data"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            credit_score_obj, created = CreditScore.objects.get_or_create(
                farmer=request.user,
                defaults={'score': 300}
            )

            # Get eligible loan offers
            loan_offers = LoanOffer.objects.filter(
                is_active=True,
                min_credit_score__lte=credit_score_obj.score
            ).order_by('interest_rate')[:5]

            credit_data = {
                'score': credit_score_obj.score,
                'score_label': credit_score_obj.score_label,
                'score_color': credit_score_obj.score_color,
                'payment_reliability': float(credit_score_obj.payment_reliability),
                'sales_history_months': credit_score_obj.sales_history_months,
                'total_revenue': float(credit_score_obj.total_revenue),
                'loan_offers': LoanOfferSerializer(loan_offers, many=True).data
            }

            serializer = CreditOverviewSerializer(data=credit_data)
            if serializer.is_valid():
                return Response(serializer.data)
            else:
                return Response(serializer.errors, status=400)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def calculate_score(self, request):
        """Calculate and update credit score based on performance"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            # Calculate credit score based on actual performance
            credit_score_obj = AnalyticsCalculationService.calculate_credit_score(request.user)
            
            if not credit_score_obj:
                return Response({'error': 'Failed to calculate credit score'}, status=500)

            return Response({
                'message': 'Credit score calculated successfully',
                'new_score': credit_score_obj.score,
                'score_label': credit_score_obj.score_label,
                'payment_reliability': credit_score_obj.payment_reliability,
                'sales_history_months': credit_score_obj.sales_history_months,
                'total_revenue': credit_score_obj.total_revenue
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LoanOfferViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for loan offers"""
    queryset = LoanOffer.objects.filter(is_active=True)
    serializer_class = LoanOfferSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['interest_rate']

    @action(detail=False, methods=['get'])
    def for_farmer(self, request):
        """Get loan offers available for the authenticated farmer"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            # Get eligible loan offers based on actual performance
            eligible_offers = AnalyticsCalculationService.get_eligible_loan_offers(request.user)

            return Response(LoanOfferSerializer(eligible_offers, many=True).data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class MonthlyReportViewSet(viewsets.ModelViewSet):
    """ViewSet for monthly reports"""
    serializer_class = MonthlyReportSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['-generated_at']

    def get_queryset(self):
        """Return reports for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return MonthlyReport.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return MonthlyReport.objects.all()
        return MonthlyReport.objects.none()

    @action(detail=False, methods=['get'])
    def overview(self, request):
        """Get reports overview data"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            monthly_reports = MonthlyReport.objects.filter(
                farmer=request.user
            ).order_by('-generated_at')[:5]

            report_types = [
                'Monthly Sales Report',
                'Quarterly Performance',
                'Annual Summary',
                'Product Analysis'
            ]

            periods = [
                'March 2024',
                'February 2024',
                'January 2024',
                'Q1 2024'
            ]

            report_data = {
                'monthly_reports': MonthlyReportSerializer(monthly_reports, many=True).data,
                'report_types': report_types,
                'periods': periods
            }

            serializer = ReportOverviewSerializer(data=report_data)
            if serializer.is_valid():
                return Response(serializer.data)
            else:
                return Response(serializer.errors, status=400)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def generate_report(self, request):
        """Generate a new monthly report"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            report_type = request.data.get('report_type', 'Monthly Sales Report')
            period = request.data.get('period', 'March 2024')

            # Check if report already exists for this period
            existing_report = MonthlyReport.objects.filter(
                farmer=request.user,
                period=period
            ).first()

            if existing_report:
                # Update existing report
                today = timezone.now().date()
                last_30_days = today - timedelta(days=30)

                farmer_orders = Order.objects.filter(
                    items__product__farmer=request.user,
                    status='delivered',
                    created_at__gte=last_30_days
                )

                revenue = farmer_orders.aggregate(total=Sum('total_amount'))['total'] or 0
                sales = farmer_orders.count()
                profit = revenue * 0.3  # Simplified profit calculation (30% margin)
                profit_margin = 30.0
                top_category = "Vegetables"
                growth_rate = 15.2

                # Update existing report
                existing_report.revenue = revenue
                existing_report.sales = sales
                existing_report.profit = profit
                existing_report.profit_margin = profit_margin
                existing_report.top_category = top_category
                existing_report.growth_rate = growth_rate
                existing_report.report_data = {
                    'report_type': report_type,
                    'generated_at': timezone.now().isoformat()
                }
                existing_report.save()

                return Response({
                    'message': 'Report updated successfully',
                    'report': MonthlyReportSerializer(existing_report).data
                })

            # Calculate report data for new report
            today = timezone.now().date()
            last_30_days = today - timedelta(days=30)

            farmer_orders = Order.objects.filter(
                items__product__farmer=request.user,
                status='delivered',
                created_at__gte=last_30_days
            )

            revenue = farmer_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            sales = farmer_orders.count()
            profit = revenue * 0.3  # Simplified profit calculation (30% margin)
            profit_margin = 30.0
            top_category = "Vegetables"
            growth_rate = 15.2

            # Create new report
            report = MonthlyReport.objects.create(
                farmer=request.user,
                period=period,
                revenue=revenue,
                sales=sales,
                profit=profit,
                profit_margin=profit_margin,
                top_category=top_category,
                growth_rate=growth_rate,
                report_data={
                    'report_type': report_type,
                    'generated_at': timezone.now().isoformat()
                }
            )

            return Response({
                'message': 'Report generated successfully',
                'report': MonthlyReportSerializer(report).data
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download a monthly report in various formats"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            report = self.get_object()
            format_type = request.query_params.get('format', 'pdf').lower()
            
            # Generate report data
            report_data = {
                'period': report.period,
                'revenue': float(report.revenue),
                'sales': report.sales,
                'profit': float(report.profit),
                'profit_margin': float(report.profit_margin),
                'top_category': report.top_category,
                'growth_rate': float(report.growth_rate),
                'generated_at': report.generated_at.strftime('%Y-%m-%d %H:%M:%S'),
                'farmer_name': f"{request.user.first_name} {request.user.last_name}",
                'farmer_email': request.user.email
            }

            if format_type == 'json':
                from django.http import JsonResponse
                return JsonResponse(report_data, safe=False)
            
            elif format_type == 'csv':
                import csv
                from django.http import HttpResponse
                
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="monthly_report_{report.period.replace(" ", "_")}.csv"'
                
                writer = csv.writer(response)
                writer.writerow(['Field', 'Value'])
                for key, value in report_data.items():
                    writer.writerow([key.replace('_', ' ').title(), value])
                
                return response
            
            elif format_type == 'excel':
                try:
                    import openpyxl
                    from django.http import HttpResponse
                    
                    # Create Excel workbook
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Monthly Report"
                    
                    # Add headers
                    ws['A1'] = 'Field'
                    ws['B1'] = 'Value'
                    
                    # Add data
                    row = 2
                    for key, value in report_data.items():
                        ws[f'A{row}'] = key.replace('_', ' ').title()
                        ws[f'B{row}'] = value
                        row += 1
                    
                    # Style headers
                    from openpyxl.styles import Font
                    header_font = Font(bold=True)
                    ws['A1'].font = header_font
                    ws['B1'].font = header_font
                    
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="monthly_report_{report.period.replace(" ", "_")}.xlsx"'
                    
                    wb.save(response)
                    return response
                    
                except ImportError:
                    return Response({'error': 'Excel export requires openpyxl package'}, status=500)
            
            elif format_type == 'pdf':  # Explicit PDF format
                try:
                    from reportlab.pdfgen import canvas
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib.units import inch
                    from django.http import HttpResponse
                    
                    response = HttpResponse(content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="monthly_report_{report.period.replace(" ", "_")}.pdf"'
                    
                    # Create PDF
                    p = canvas.Canvas(response, pagesize=letter)
                    width, height = letter
                    
                    # Title
                    p.setFont("Helvetica-Bold", 16)
                    p.drawString(width/2 - 100, height - 50, f"Monthly Report - {report.period}")
                    
                    # Content
                    p.setFont("Helvetica", 12)
                    y_position = height - 100
                    
                    for key, value in report_data.items():
                        if key != 'period':
                            label = key.replace('_', ' ').title()
                            p.drawString(50, y_position, f"{label}: {value}")
                            y_position -= 20
                    
                    p.showPage()
                    p.save()
                    return response
                    
                except ImportError:
                    return Response({'error': 'PDF export requires reportlab package'}, status=500)
            
            else:  # Default to PDF (fallback)
                try:
                    from reportlab.pdfgen import canvas
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib.units import inch
                    from django.http import HttpResponse
                    
                    response = HttpResponse(content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="monthly_report_{report.period.replace(" ", "_")}.pdf"'
                    
                    # Create PDF
                    p = canvas.Canvas(response, pagesize=letter)
                    width, height = letter
                    
                    # Title
                    p.setFont("Helvetica-Bold", 16)
                    p.drawString(width/2 - 100, height - 50, f"Monthly Report - {report.period}")
                    
                    # Content
                    p.setFont("Helvetica", 12)
                    y_position = height - 100
                    
                    for key, value in report_data.items():
                        if key != 'period':  # Skip period as it's in title
                            label = key.replace('_', ' ').title()
                            p.drawString(50, y_position, f"{label}: {value}")
                            y_position -= 20
                    
                    p.showPage()
                    p.save()
                    return response
                    
                except ImportError:
                    return Response({'error': 'PDF export requires reportlab package'}, status=500)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ProductPerformanceViewSet(viewsets.ModelViewSet):
    """ViewSet for product performance tracking"""
    serializer_class = ProductPerformanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['-date']

    def get_queryset(self):
        """Return performance data for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return ProductPerformance.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return ProductPerformance.objects.all()
        return ProductPerformance.objects.none()


# Quick Actions Viewsets
class SalesTargetViewSet(viewsets.ModelViewSet):
    """ViewSet for sales targets management"""
    serializer_class = SalesTargetSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['-created_at']

    def get_queryset(self):
        """Return targets for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return SalesTarget.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return SalesTarget.objects.all()
        return SalesTarget.objects.none()

    def perform_create(self, serializer):
        """Set the farmer when creating a target"""
        serializer.save(farmer=self.request.user)

    @action(detail=True, methods=['post'])
    def update_progress(self, request, pk=None):
        """Update target progress based on current performance"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            target = self.get_object()
            
            # Calculate current value based on target type
            today = timezone.now().date()
            last_30_days = today - timedelta(days=30)
            
            farmer_orders = Order.objects.filter(
                items__product__farmer=request.user,
                status='delivered',
                created_at__gte=last_30_days
            )

            if target.target_type == 'revenue':
                current_value = farmer_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            elif target.target_type == 'sales':
                current_value = farmer_orders.count()
            elif target.target_type == 'growth':
                # Calculate growth percentage
                current_month_revenue = farmer_orders.aggregate(total=Sum('total_amount'))['total'] or 0
                previous_month = last_30_days - timedelta(days=30)
                previous_month_revenue = Order.objects.filter(
                    items__product__farmer=request.user,
                    status='delivered',
                    created_at__gte=previous_month,
                    created_at__lt=last_30_days
                ).aggregate(total=Sum('total_amount'))['total'] or 0
                
                if previous_month_revenue > 0:
                    current_value = ((current_month_revenue - previous_month_revenue) / previous_month_revenue) * 100
                else:
                    current_value = 0
            else:
                current_value = 0

            target.current_value = current_value
            target.calculate_progress()
            target.save()

            return Response({
                'message': 'Target progress updated successfully',
                'current_value': float(target.current_value),
                'progress_percentage': float(target.progress_percentage),
                'is_completed': target.is_completed
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class PaymentAnalysisViewSet(viewsets.ModelViewSet):
    """ViewSet for payment analysis"""
    serializer_class = PaymentAnalysisSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['-generated_at']

    def get_queryset(self):
        """Return analyses for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return PaymentAnalysis.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return PaymentAnalysis.objects.all()
        return PaymentAnalysis.objects.none()

    @action(detail=False, methods=['post'])
    def generate_analysis(self, request):
        """Generate payment analysis for the farmer"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            analysis_period = request.data.get('analysis_period', 'March 2024')
            
            # Calculate analysis period dates
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            end_date = today

            # Get farmer's orders
            farmer_orders = Order.objects.filter(
                items__product__farmer=request.user,
                created_at__gte=start_date,
                created_at__lte=end_date
            )

            # Calculate payment metrics
            total_payments = farmer_orders.filter(status='delivered').aggregate(
                total=Sum('total_amount')
            )['total'] or 0

            pending_payments = farmer_orders.filter(status='pending').aggregate(
                total=Sum('total_amount')
            )['total'] or 0

            overdue_payments = farmer_orders.filter(
                status='pending',
                created_at__lt=today - timedelta(days=7)
            ).aggregate(total=Sum('total_amount'))['total'] or 0

            # Calculate payment reliability
            total_orders = farmer_orders.count()
            delivered_orders = farmer_orders.filter(status='delivered').count()
            
            if total_orders > 0:
                payment_reliability_score = (delivered_orders / total_orders) * 100
            else:
                payment_reliability_score = 0

            # Create analysis
            analysis = PaymentAnalysis.objects.create(
                farmer=request.user,
                analysis_period=analysis_period,
                start_date=start_date,
                end_date=end_date,
                total_payments=total_payments,
                pending_payments=pending_payments,
                overdue_payments=overdue_payments,
                on_time_payments=delivered_orders,
                late_payments=total_orders - delivered_orders,
                payment_reliability_score=payment_reliability_score,
                is_completed=True,
                payment_trends=[
                    {'date': 'Week 1', 'payments': float(total_payments * 0.3)},
                    {'date': 'Week 2', 'payments': float(total_payments * 0.25)},
                    {'date': 'Week 3', 'payments': float(total_payments * 0.25)},
                    {'date': 'Week 4', 'payments': float(total_payments * 0.2)}
                ],
                customer_payment_behavior={
                    'on_time_percentage': float(payment_reliability_score),
                    'average_payment_time': 3.5,
                    'preferred_payment_method': 'Mobile Money'
                }
            )

            return Response({
                'message': 'Payment analysis generated successfully',
                'analysis': PaymentAnalysisSerializer(analysis).data
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ExportRequestViewSet(viewsets.ModelViewSet):
    """ViewSet for export requests"""
    serializer_class = ExportRequestSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ['-requested_at']

    def get_queryset(self):
        """Return export requests for the authenticated farmer only"""
        if self.request.user.is_farmer:
            return ExportRequest.objects.filter(farmer=self.request.user)
        elif self.request.user.is_admin:
            return ExportRequest.objects.all()
        return ExportRequest.objects.none()

    def perform_create(self, serializer):
        """Set the farmer when creating an export request"""
        serializer.save(farmer=self.request.user)

    @action(detail=True, methods=['post'])
    def process_export(self, request, pk=None):
        """Process the export request and generate file"""
        if not request.user.is_farmer:
            return Response({'error': 'Only farmers can access this endpoint'}, status=403)

        try:
            export_request = self.get_object()
            
            # Simulate export processing
            export_request.status = 'processing'
            export_request.save()

            # Calculate date range
            today = timezone.now().date()
            if export_request.date_range == 'last_month':
                start_date = today - timedelta(days=30)
                end_date = today
            elif export_request.date_range == 'last_quarter':
                start_date = today - timedelta(days=90)
                end_date = today
            elif export_request.date_range == 'last_year':
                start_date = today - timedelta(days=365)
                end_date = today
            else:
                start_date = export_request.start_date
                end_date = export_request.end_date

            # Get data based on export type
            farmer_orders = Order.objects.filter(
                items__product__farmer=request.user,
                created_at__gte=start_date,
                created_at__lte=end_date
            )

            if export_request.export_type == 'sales_data':
                export_data = list(farmer_orders.values(
                    'id', 'total_amount', 'status', 'created_at'
                ))
            elif export_request.export_type == 'revenue_report':
                export_data = {
                    'total_revenue': float(farmer_orders.aggregate(
                        total=Sum('total_amount')
                    )['total'] or 0),
                    'total_orders': farmer_orders.count(),
                    'period': f"{start_date} to {end_date}"
                }
            elif export_request.export_type == 'product_performance':
                export_data = list(OrderItem.objects.filter(
                    order__in=farmer_orders
                ).values('product__name').annotate(
                    total_sales=Sum('quantity'),
                    total_revenue=Sum(F('quantity') * F('unit_price'))
                ))
            else:
                export_data = {'message': 'Export data generated'}

            # Update export request
            export_request.status = 'completed'
            export_request.completed_at = timezone.now()
            export_request.export_data = export_data
            export_request.file_url = f"/exports/{export_request.id}.{export_request.export_format}"
            export_request.file_size = len(str(export_data)) * 8  # Approximate size
            export_request.save()

            return Response({
                'message': 'Export completed successfully',
                'file_url': export_request.file_url,
                'file_size': export_request.file_size,
                'export_data': export_data
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)
